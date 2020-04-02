# -*- coding: utf-8 -*-
import json
import time

from openpyxl.styles import Font, Side, Border, Alignment

from HousePrice.items import ErShouFangVillageItem, ErShouFangCityItem, ErShouFangSourceItem, RentingHouseSourceItem, NewHouseItem
from openpyxl import Workbook
# 设定excel的样式
font = Font(name='宋体', size=11, bold=False, vertAlign=None, underline='none', strike=False, color='000000')
side = Side(border_style='thin', color='000000')
border = Border(left=side, right=side, top=side, bottom=side)
alignment_enter = Alignment(horizontal='center', vertical='center')

class HousepriceExcelPipeline(object):
    def __init__(self):
        self.city_items = []
        self.village_items = []
        self.source_items = []
        self.renting_house_items = []
        self.new_house_items = []

    def open_spider(self, spider):
        print('open spider')
        self.wb = Workbook()
        self.wb.remove(self.wb.active)

    def adjust_column_width(self, ws):
        """
        自适应列的宽度
        :param ws:
        :return:
        """
        for col in ws.columns:
            max_length = 0
            column = col[0].column
            unmerged_cells = list(
                filter(lambda cell_to_check: cell_to_check.coordinate not in ws.merged_cells, col))
            for cell in col:
                try:
                    if cell.coordinate in ws.merged_cells:  # not check merge_cells
                        continue
                    value_length = len(str(cell.value))
                    if value_length > max_length:
                        max_length = value_length
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[unmerged_cells[0].column_letter].width = adjusted_width

    def close_spider(self, spider):
        ws1 = self.wb.create_sheet(title='二手房 1.1 城市区域信息', index=0)
        ws1_header = ['城市', '所属区', '挂盘数量', '均价（=本版块所有在售房源价格的平均价）', '价格增幅（=本月均价-上月均价/上月均价*100%）']
        ws1_row_index = 1
        for i, value in enumerate(ws1_header):
            ws1.cell(row=ws1_row_index, column=i + 1, value=value)
        ws1_row_index += 1
        for city_item in self.city_items:
            ws1.cell(row=ws1_row_index, column=1, value=city_item['city'])
            ws1.cell(row=ws1_row_index, column=2, value=city_item['area'])
            ws1.cell(row=ws1_row_index, column=3, value=city_item['guapan_num'])
            ws1.cell(row=ws1_row_index, column=4, value=city_item['avg_price'])
            ws1.cell(row=ws1_row_index, column=5, value=city_item['price_width'])
            ws1_row_index += 1

        ws2 = self.wb.create_sheet(title='二手房 1.2 小区信息', index=1)
        ws2_header = ['城市', '小区所属区', '板块', '小区名称', '挂盘数量', '挂盘数量涨幅（=本月挂盘数量-上月挂盘数量）/上月挂盘数量*100%）', '小区均价（元）',
                      '小区均价较上周涨幅=（本月小区均价-上月小区均价）/上月小区均价*100%）', '小区均价链接']
        ws2_row_index = 1
        for i, value in enumerate(ws2_header):
            ws2.cell(row=ws2_row_index, column=i + 1, value=value)
        ws2_row_index += 1
        for village_item in self.village_items:
            ws2.cell(row=ws2_row_index, column=1, value=village_item['city'])
            ws2.cell(row=ws2_row_index, column=2, value=village_item['area'])
            ws2.cell(row=ws2_row_index, column=3, value=village_item['bankuai_name'])
            ws2.cell(row=ws2_row_index, column=4, value=village_item['name'])
            ws2.cell(row=ws2_row_index, column=5, value=village_item['guapan_num'])
            ws2.cell(row=ws2_row_index, column=7, value=village_item['village_avg_price'])
            ws2.cell(row=ws2_row_index, column=9, value=village_item['village_link'])
            ws2_row_index += 1

        ws3 = self.wb.create_sheet(title='二手房 1.3 房源信息', index=2)
        ws3_header = ['城市', '城区', '商圈', '小区名', '挂牌时间', '朝向', '居/室/厅', '楼层', '建筑面积', '套内面积', '总价 （万元）', '单价 （元）',
                      '小区均价 （元）', '小区均价链接',
                      '建筑时间', '小区笋盘幅度（=房源单价/小区均价*100%）', '房源涨价幅度（=（本房源本周单价-本房源两周前单价）/本房源两周前单价*100%）', '房源链接']
        ws3_row_index = 1
        for i, value in enumerate(ws3_header):
            ws3.cell(row=ws3_row_index, column=i + 1, value=value)
        ws3_row_index += 1
        for source_item in self.source_items:
            ws3.cell(row=ws3_row_index, column=1, value=source_item['city'])
            ws3.cell(row=ws3_row_index, column=2, value=source_item['area'])
            ws3.cell(row=ws3_row_index, column=3, value=source_item['business_circle'])
            ws3.cell(row=ws3_row_index, column=4, value=source_item['village_name'])
            ws3.cell(row=ws3_row_index, column=5, value=source_item['listing_time'])
            ws3.cell(row=ws3_row_index, column=6, value=source_item['orientation'])
            ws3.cell(row=ws3_row_index, column=7, value=source_item['residence_room'])
            ws3.cell(row=ws3_row_index, column=8, value=source_item['floor'])
            ws3.cell(row=ws3_row_index, column=9, value=source_item['area1'])
            ws3.cell(row=ws3_row_index, column=10, value=source_item['area2'])
            ws3.cell(row=ws3_row_index, column=11, value=source_item['all_price'])
            ws3.cell(row=ws3_row_index, column=12, value=source_item['single_price'])
            ws3.cell(row=ws3_row_index, column=13, value=source_item['village_avg_price'])
            ws3.cell(row=ws3_row_index, column=14, value=source_item['village_link'])
            ws3.cell(row=ws3_row_index, column=15, value=source_item['build_time'])
            ws3.cell(row=ws3_row_index, column=18, value=source_item['link'])
            ws3_row_index += 1

        ws4 = self.wb.create_sheet(title='租房 2.1 房源信息', index=3)
        ws4_row_index = 1
        ws4_header = ['城市', '城区', '板块', '小区名', '朝向', '居/室/厅', '楼层', '面积', '租金', '租期', '房源链接']
        for i, value in enumerate(ws4_header):
            ws4.cell(row=ws4_row_index, column=i + 1, value=value)
        ws4_row_index += 1
        for renting_house_item in self.renting_house_items:
            ws4.cell(row=ws4_row_index, column=1, value=renting_house_item['city'])
            ws4.cell(row=ws4_row_index, column=2, value=renting_house_item['area'])
            ws4.cell(row=ws4_row_index, column=3, value=renting_house_item['business_circle'])
            ws4.cell(row=ws4_row_index, column=4, value=renting_house_item['village_name'])
            ws4.cell(row=ws4_row_index, column=5, value=renting_house_item['orientation'])
            ws4.cell(row=ws4_row_index, column=6, value=renting_house_item['residence_room'])
            ws4.cell(row=ws4_row_index, column=7, value=renting_house_item['floor'])
            ws4.cell(row=ws4_row_index, column=8, value=renting_house_item['house_area'])
            ws4.cell(row=ws4_row_index, column=9, value=renting_house_item['price'])
            ws4.cell(row=ws4_row_index, column=10, value=renting_house_item['renting_time'])
            ws4.cell(row=ws4_row_index, column=11, value=renting_house_item['link'])
            ws4_row_index += 1

        ws5 = self.wb.create_sheet(title='新房 3.1 楼盘信息', index=4)
        ws5_row_index = 1
        ws5_header = ['城市', '新盘所属区', '新盘名称', '别名', '最新开盘时间', '新盘单价（元）（均价）', '新盘总价（万元）（均价）', '新盘链接']
        for i, value in enumerate(ws5_header):
            ws5.cell(row=ws5_row_index, column=i + 1, value=value)
        ws5_row_index += 1
        for new_house_item in self.new_house_items:
            ws5.cell(row=ws5_row_index, column=1, value=new_house_item['city'])
            ws5.cell(row=ws5_row_index, column=2, value=new_house_item['area'])
            ws5.cell(row=ws5_row_index, column=3, value=new_house_item['name'])
            ws5.cell(row=ws5_row_index, column=4, value=new_house_item['another_name'])
            ws5.cell(row=ws5_row_index, column=5, value=new_house_item['new_kp_time'])
            ws5.cell(row=ws5_row_index, column=6, value=new_house_item['new_kp_single_price'])
            ws5.cell(row=ws5_row_index, column=7, value=new_house_item['new_kp_all_price'])
            ws5.cell(row=ws5_row_index, column=8, value=new_house_item['link'])
            ws5_row_index += 1


        self.adjust_column_width(ws1)
        self.adjust_column_width(ws2)
        self.adjust_column_width(ws3)
        self.adjust_column_width(ws4)
        self.adjust_column_width(ws5)

        self.wb.save(filename='房价数据_{}_{}.xlsx'.format(spider.name, int(time.time())))
        spider.logger.info('close spider: {}'.format(spider.name))

    def process_item(self, item, spider):
        if isinstance(item, ErShouFangCityItem):
            return self.handle_city_item(item)
        elif isinstance(item, ErShouFangVillageItem):
            return self.handle_village_item(item)
        elif isinstance(item, ErShouFangSourceItem):
            return self.handle_house_source_item(item)
        elif isinstance(item, RentingHouseSourceItem):
            return self.handle_renting_house_item(item)
        elif isinstance(item, NewHouseItem):
            return self.handle_new_house_item(item)
    def handle_city_item(self, item):
        self.city_items.append(dict(item))
        return item

    def handle_village_item(self, item):
        self.village_items.append(dict(item))
        return item

    def handle_house_source_item(self, item):
        self.source_items.append(dict(item))
        return item

    def handle_renting_house_item(self, item):
        self.renting_house_items.append(dict(item))
        return item

    def handle_new_house_item(self, item):
        self.new_house_items.append(dict(item))
        return item